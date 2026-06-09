from __future__ import annotations

import hashlib
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st

from src.datas import agora_brasilia, datetime_arquivo_brasilia
from src.persistencia import carregar_bytes_detalhado, salvar_bytes
from src.tratamento import (
    COLUNAS_ACOES,
    COLUNAS_ACOES_BASE,
    COLUNAS_BUSSOLA,
    COLUNAS_CONTATO,
    COLUNAS_PAINEL,
    COLUNAS_PRODUTOS_MIX,
    padronizar_colunas,
    preparar_acoes,
    preparar_base_vendas,
    preparar_painel_equipe,
    preparar_produtos_mix,
    serie_data,
    validar_colunas_esperadas,
)


ROOT_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT_DIR / "data"

ARQUIVOS_PADRAO = {
    "bussola": DATA_DIR / "bussola.xlsx",
    "painel": DATA_DIR / "PAINEL EQUIPE NORTE.xlsx",
    "acoes": DATA_DIR / "template_acoes_promocionais.xlsx",
    "produtos_mix": DATA_DIR / "template_produtos_mix.xlsx",
}

ABAS_PADRAO = {
    "bussola": "Pedidos",
    "painel": "Planilha1",
    "acoes": 0,
    "produtos_mix": 0,
}


def _uploads_sessao() -> dict[str, dict[str, object]]:
    return st.session_state.setdefault("uploads_bases", {})


def _formatar_mb(tamanho_bytes: int) -> str:
    return f"{tamanho_bytes / (1024 * 1024):.2f} MB"


def _cnpj_valido(cnpj: object) -> bool:
    texto = str(cnpj or "").strip()
    return bool(texto and texto.isdigit() and len(texto) == 14 and texto != "0" * 14 and len(set(texto)) > 1)


def _celula_vazia(valor: object) -> bool:
    if valor is None:
        return True
    if pd.isna(valor):
        return True
    return str(valor).strip().lower() in {"", "nan", "none", "<na>", "nat", "null"}


def _colunas_unicas(cabecalho: list[object]) -> list[str]:
    nomes: list[str] = []
    contagem: dict[str, int] = {}
    for idx, valor in enumerate(cabecalho):
        nome = str(valor).strip() if not _celula_vazia(valor) else f"Unnamed: {idx}"
        ocorrencias = contagem.get(nome, 0)
        contagem[nome] = ocorrencias + 1
        nomes.append(nome if ocorrencias == 0 else f"{nome}.{ocorrencias}")
    return nomes


def _ler_painel_upload(conteudo: bytes) -> pd.DataFrame:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
        ws = wb[ABAS_PADRAO["painel"]] if ABAS_PADRAO["painel"] in wb.sheetnames else wb.worksheets[0]
        cabecalho: list[str] | None = None
        linhas: list[list[object]] = []
        vazias_consecutivas = 0
        for linha in ws.iter_rows(values_only=True):
            valores = list(linha)
            if all(_celula_vazia(valor) for valor in valores):
                if cabecalho is not None:
                    vazias_consecutivas += 1
                    if vazias_consecutivas >= 500:
                        break
                continue
            vazias_consecutivas = 0
            if cabecalho is None:
                cabecalho = _colunas_unicas(valores)
                continue
            if len(valores) < len(cabecalho):
                valores.extend([""] * (len(cabecalho) - len(valores)))
            linhas.append(valores[: len(cabecalho)])
        wb.close()
        return pd.DataFrame(linhas, columns=cabecalho or [])
    except Exception:
        try:
            return pd.read_excel(BytesIO(conteudo), sheet_name=ABAS_PADRAO["painel"], dtype=str, engine="openpyxl")
        except ValueError:
            return pd.read_excel(BytesIO(conteudo), sheet_name=0, dtype=str, engine="openpyxl")


def _linha_util(serie: pd.Series) -> bool:
    for valor in serie:
        if pd.notna(valor) and str(valor).strip().lower() not in {"", "nan", "none", "<na>", "nat", "null"}:
            return True
    return False


def _compactar_upload_painel(conteudo: bytes) -> tuple[bytes, int]:
    painel_raw = _ler_painel_upload(conteudo)
    painel = preparar_painel_equipe(painel_raw)
    if painel.empty:
        compacto = pd.DataFrame(columns=COLUNAS_PAINEL + COLUNAS_CONTATO + ["cnpj_limpo", "grupo_sip", "cliente_ativo"])
    else:
        colunas_usuario = [col for col in COLUNAS_PAINEL + COLUNAS_CONTATO if col in painel.columns]
        painel = painel[painel[colunas_usuario].apply(_linha_util, axis=1)]
        painel = painel[painel["cnpj_limpo"].apply(_cnpj_valido)]
        painel = painel.drop_duplicates("cnpj_limpo", keep="first").reset_index(drop=True)
        colunas_tecnicas = ["cnpj_limpo", "grupo_sip", "cliente_ativo"]
        colunas_compactas = [
            col
            for col in COLUNAS_PAINEL + COLUNAS_CONTATO + colunas_tecnicas
            if col in painel.columns
        ]
        compacto = painel[colunas_compactas].copy()

    saida = BytesIO()
    with pd.ExcelWriter(saida, engine="openpyxl") as writer:
        compacto.to_excel(writer, sheet_name="Planilha1", index=False)
    return saida.getvalue(), int(len(compacto.index))


def _criar_backup_upload(chave: str) -> None:
    fonte_atual = _fonte_base_excel(chave, exibir_alertas=False)
    conteudo = fonte_atual.get("conteudo")
    if not isinstance(conteudo, bytes) or not conteudo:
        return
    destino_dir = DATA_DIR / "_backups_uploads" / chave
    destino_dir.mkdir(parents=True, exist_ok=True)
    timestamp = agora_brasilia().strftime("%Y%m%d_%H%M%S")
    destino = destino_dir / f"{timestamp}_{ARQUIVOS_PADRAO[chave].name}"
    destino.write_bytes(conteudo)


def registrar_upload(chave: str, arquivo) -> dict[str, object] | None:
    if arquivo is None:
        return None
    conteudo = arquivo.getvalue()
    tamanho_original = len(conteudo)

    try:
        _criar_backup_upload(chave)
    except Exception:
        st.warning("Não foi possível criar backup automático, mas vou tentar salvar a nova base.")

    clientes_validos: int | None = None
    if chave == "painel":
        conteudo, clientes_validos = _compactar_upload_painel(conteudo)

    _uploads_sessao()[chave] = {"name": arquivo.name, "bytes": conteudo, "updated_at": agora_brasilia().isoformat()}
    github_persistiu = salvar_bytes(chave, conteudo, f"Atualiza base {chave} pelo painel")
    st.cache_data.clear()
    mensagem = f"Upload salvo: {arquivo.name}"
    if chave == "painel" and clientes_validos is not None:
        mensagem = (
            f"Base de clientes tratada e salva: {clientes_validos:,} clientes válidos. "
            f"Tamanho reduzido de {_formatar_mb(tamanho_original)} para {_formatar_mb(len(conteudo))}."
        ).replace(",", ".")
        st.success(mensagem)
    return {
        "chave": chave,
        "arquivo": arquivo.name,
        "bytes": len(conteudo),
        "github_persistiu": github_persistiu,
        "mensagem": mensagem,
        "clientes_validos": clientes_validos,
    }


def limpar_upload_sessao(chave: str) -> None:
    _uploads_sessao().pop(chave, None)


def limpar_uploads() -> None:
    st.session_state["uploads_bases"] = {}
    st.cache_data.clear()


def _hash_simples(conteudo: bytes | None) -> str:
    return hashlib.sha256(conteudo).hexdigest()[:12] if conteudo else "-"


def _rotulo_origem(chave: str, origem: str, nome: str = "") -> str:
    if origem == "upload_sessao":
        return f"upload sessão: {nome}" if nome else "upload sessão"
    if origem == "github":
        return "GitHub persistido"
    if origem == "local_persistencia":
        return "persistência local temporária"
    if origem == "local_data":
        return f"local data/{ARQUIVOS_PADRAO[chave].name}"
    return "não encontrado"


def _alertar_fallback_bussola(origem: str) -> None:
    if origem in {"local_persistencia", "local_data"}:
        st.warning("Atenção: usando fallback local/antigo do Bússola.")


def _fonte_base_excel(chave: str, exibir_alertas: bool = True) -> dict[str, Any]:
    upload = _uploads_sessao().get(chave)
    conteudo_upload = upload.get("bytes") if upload else None
    if isinstance(conteudo_upload, bytes) and conteudo_upload:
        nome = str(upload.get("name", "")) if upload else ""
        return {
            "origem_codigo": "upload_sessao",
            "origem": _rotulo_origem(chave, "upload_sessao", nome),
            "conteudo": conteudo_upload,
            "hash_conteudo": _hash_simples(conteudo_upload),
            "atualizado_em": upload.get("updated_at") if upload else None,
        }

    persistido = carregar_bytes_detalhado(chave)
    conteudo_persistido = persistido.get("conteudo")
    origem_persistida = str(persistido.get("origem") or "nao_encontrado")
    if isinstance(conteudo_persistido, bytes) and conteudo_persistido:
        if chave == "bussola" and exibir_alertas:
            _alertar_fallback_bussola(origem_persistida)
        return {
            "origem_codigo": origem_persistida,
            "origem": _rotulo_origem(chave, origem_persistida),
            "conteudo": conteudo_persistido,
            "hash_conteudo": _hash_simples(conteudo_persistido),
            "atualizado_em": persistido.get("atualizado_em"),
        }

    caminho = ARQUIVOS_PADRAO[chave]
    if caminho.exists():
        conteudo_local = caminho.read_bytes()
        if chave == "bussola" and exibir_alertas:
            _alertar_fallback_bussola("local_data")
        return {
            "origem_codigo": "local_data",
            "origem": _rotulo_origem(chave, "local_data"),
            "conteudo": conteudo_local,
            "hash_conteudo": _hash_simples(conteudo_local),
            "atualizado_em": datetime_arquivo_brasilia(caminho),
        }

    if chave == "bussola" and exibir_alertas and persistido.get("github_ativo"):
        st.warning("Bússola não encontrado na persistência GitHub nem em fallback local.")
    return {
        "origem_codigo": "nao_encontrado",
        "origem": _rotulo_origem(chave, "nao_encontrado"),
        "conteudo": None,
        "hash_conteudo": "-",
        "atualizado_em": None,
    }


def fonte_ativa(chave: str) -> str:
    return str(_fonte_base_excel(chave, exibir_alertas=False)["origem"])


@st.cache_data(show_spinner=False)
def _ler_excel_bytes(conteudo: bytes, sheet_name: str | int) -> pd.DataFrame:
    return pd.read_excel(BytesIO(conteudo), sheet_name=sheet_name, dtype=str, engine="openpyxl")


@st.cache_data(show_spinner=False)
def _ler_excel_caminho(caminho: str, sheet_name: str | int, mtime: float) -> pd.DataFrame:
    return pd.read_excel(caminho, sheet_name=sheet_name, dtype=str, engine="openpyxl")


def _ler_excel_por_fonte(chave: str, fonte: dict[str, Any]) -> pd.DataFrame:
    conteudo = fonte.get("conteudo")
    if not isinstance(conteudo, bytes) or not conteudo:
        return pd.DataFrame()
    return _ler_excel_bytes(conteudo, ABAS_PADRAO[chave])


def _carregar_excel_com_fonte(chave: str) -> tuple[pd.DataFrame, dict[str, Any]]:
    fonte = _fonte_base_excel(chave)
    return _ler_excel_por_fonte(chave, fonte), fonte


def _carregar_excel(chave: str) -> pd.DataFrame:
    df, _fonte = _carregar_excel_com_fonte(chave)
    return df


def _periodo_bussola(df: pd.DataFrame) -> tuple[object | None, object | None]:
    if df is None or df.empty:
        return None, None
    base = padronizar_colunas(df)
    if "data_do_pedido" not in base.columns:
        return None, None
    datas = serie_data(base["data_do_pedido"]).dropna()
    if datas.empty:
        return None, None
    return datas.min(), datas.max()


def diagnostico_bussola(df: pd.DataFrame | None = None, fonte: dict[str, Any] | None = None) -> dict[str, object]:
    fonte_ativa_bussola = fonte or _fonte_base_excel("bussola", exibir_alertas=False)
    base = df if df is not None else _ler_excel_por_fonte("bussola", fonte_ativa_bussola)
    menor_data, maior_data = _periodo_bussola(base)
    return {
        "origem": fonte_ativa_bussola["origem"],
        "origem_codigo": fonte_ativa_bussola["origem_codigo"],
        "total_linhas": int(len(base.index)) if base is not None else 0,
        "menor_data_do_pedido": menor_data,
        "maior_data_do_pedido": maior_data,
        "hash_conteudo": fonte_ativa_bussola["hash_conteudo"],
        "ultima_atualizacao_usada": fonte_ativa_bussola["atualizado_em"],
    }


def carregar_bussola() -> pd.DataFrame:
    return _carregar_excel("bussola")


def carregar_painel_equipe() -> pd.DataFrame:
    return _carregar_excel("painel")


def carregar_acoes() -> pd.DataFrame:
    return _carregar_excel("acoes")


def carregar_produtos_mix() -> pd.DataFrame:
    return _carregar_excel("produtos_mix")


def carregar_dados_tratados() -> dict[str, object]:
    bussola_raw, bussola_fonte = _carregar_excel_com_fonte("bussola")
    bussola_diag = diagnostico_bussola(bussola_raw, bussola_fonte)
    painel_raw = carregar_painel_equipe()
    acoes_raw = carregar_acoes()
    produtos_raw = carregar_produtos_mix()

    avisos: list[str] = []
    avisos.extend(validar_colunas_esperadas(bussola_raw, COLUNAS_BUSSOLA, "bussola.xlsx"))
    avisos.extend(validar_colunas_esperadas(painel_raw, COLUNAS_PAINEL, "PAINEL EQUIPE NORTE.xlsx"))
    if acoes_raw.empty:
        avisos.append("Foco semanal: sem ações cadastradas. Use a tela Importar Bases para baixar o modelo.")
    if produtos_raw.empty:
        avisos.append("template_produtos_mix.xlsx: sem produtos classificados. Produtos vendidos ficarão como SEM CLASSIFICACAO.")
    else:
        avisos.extend(validar_colunas_esperadas(produtos_raw, COLUNAS_PRODUTOS_MIX, "template_produtos_mix.xlsx"))
    if not acoes_raw.empty:
        avisos.extend(validar_colunas_esperadas(acoes_raw, COLUNAS_ACOES_BASE, "Foco semanal"))

    clientes = preparar_painel_equipe(painel_raw)
    produtos_mix = preparar_produtos_mix(produtos_raw)
    acoes = preparar_acoes(acoes_raw)
    vendas = preparar_base_vendas(bussola_raw, clientes, produtos_mix)

    return {
        "vendas": vendas,
        "clientes": clientes,
        "produtos_mix": produtos_mix,
        "acoes": acoes,
        "avisos": avisos,
        "raw_bussola": bussola_raw,
        "diagnostico_bussola": bussola_diag,
        "raw_painel": painel_raw,
        "raw_acoes": acoes_raw,
        "raw_produtos_mix": produtos_raw,
    }


def modelo_acoes() -> pd.DataFrame:
    return pd.DataFrame(
        columns=[
            "NOME_ACAO",
            "EAN",
            "PRODUTO",
            "TIPO_MIX",
            "DISTRIBUIDORA",
            "DESCONTO",
            "DATA_INICIO",
            "DATA_FIM",
            "CONSULTOR",
            "OBSERVACAO",
            "STATUS",
            "META_UNIDADES",
            "META_CNPJS",
            "TIPO_META_UNIDADES",
            "ESCOPO_META",
            "METAS_CONSULTORES",
        ]
    )


def modelo_produtos_mix() -> pd.DataFrame:
    return pd.DataFrame(columns=COLUNAS_PRODUTOS_MIX)
